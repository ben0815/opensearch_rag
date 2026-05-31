"""Document processing functionality."""

from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import os
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, AsyncGenerator
from zipfile import BadZipFile

import chardet
import fitz

from app.loader.chunk_splitter import ChunkSplitter, _load_tokenizer, _truncate_to_tokens
from app.loader.config import LoaderConfig
from app.metadata.redis_service import DocumentMetadata, RedisMetadataService
from app.utils.logging_config import setup_logger

if TYPE_CHECKING:
    from langchain_core.documents import Document
    from app.loader.vector_store import VectorStore

logger = setup_logger(__name__)

_MAX_PDF_PAGES = int(os.getenv("MAX_PDF_PAGES", "500"))
_MAX_TEXT_PER_PAGE = 1_000_000  # 1 MB Text pro Seite als DoS-Schutz


def _detect_pdf_header_footer(doc) -> set[str]:
    """Erkennt Kopf-/Fußzeilen im PDF.

    Text, der auf > 30 % der Seiten in den oberen oder unteren 10 % erscheint
    und kürzer als 200 Zeichen ist, gilt als Kopf-/Fußzeile und wird beim
    Extrahieren übersprungen. Mindestens 3 Seiten erforderlich.
    """
    total_pages = len(doc)
    if total_pages < 3:
        return set()
    threshold = max(2, int(total_pages * 0.3))
    counter: Counter[str] = Counter()
    for page in doc:
        page_h = page.rect.height
        for block in page.get_text("blocks"):
            _, by0, _, by1, text, *_ = block
            text = text.strip()
            if not text or len(text) > 200:
                continue
            if by1 < page_h * 0.1 or by0 > page_h * 0.9:
                counter[text] += 1
    return {text for text, count in counter.items() if count >= threshold}


class DocumentProcessor:
    """Process and store documents."""

    def __init__(
        self,
        config: LoaderConfig | None = None,
        vector_store: VectorStore | None = None,
        instance_slug: str = "default",
        redis_service: RedisMetadataService | None = None,
    ) -> None:
        self.config = config or LoaderConfig()
        if vector_store is None:
            raise ValueError('vector_store is required for DocumentProcessor')
        self.vector_store = vector_store
        self._instance_slug = instance_slug
        # Injizierte redis_service nutzen (App-Singleton), sonst Fallback für CLI-Nutzung
        self.metadata_service = redis_service or RedisMetadataService.from_config(
            self.config, instance_slug
        )
        self._tokenizer = _load_tokenizer(self.config.tokenizer_model_id)
        self._chunk_splitter = ChunkSplitter(
            chunk_size=self.config.chunk_size,
            chunk_overlap=self.config.chunk_overlap,
            embedding_max_tokens=self.config.embedding_max_tokens,
            tokenizer_model_id=self.config.tokenizer_model_id,
            min_chunk_size=self.config.min_chunk_size,
        )
        self._warnings: list[str] = []

    async def _calculate_file_hash(self, file_path: Path) -> str:
        """Calculate SHA-256 hash of file."""
        def _hash():
            sha256_hash = hashlib.sha256()
            with open(file_path, 'rb') as f:
                for byte_block in iter(lambda: f.read(4096), b''):
                    sha256_hash.update(byte_block)
            return sha256_hash.hexdigest()

        return await asyncio.to_thread(_hash)

    async def process_chunk(self, chunk: Document) -> None:
        """Embed and store a single document chunk.

        Content is truncated to embedding_max_tokens before the first attempt.
        Retries with half the token count on context-length errors — a safety
        net for tokenizer edge cases (e.g. very long numeric or special-char sequences).
        """
        store = self.vector_store.get_store()
        content = _truncate_to_tokens(chunk.page_content, self.config.embedding_max_tokens, self._tokenizer)

        while content:
            try:
                return await asyncio.to_thread(
                    store.add_texts,
                    texts=[content],
                    metadatas=[chunk.metadata],
                )
            except Exception as e:
                msg = str(e).lower()
                if 'input length' in msg or 'context length' in msg:
                    ids = self._tokenizer.encode(content, add_special_tokens=False)
                    reduced_token_count = len(ids) // 2
                    if reduced_token_count < 10:
                        logger.error('Chunk too dense to embed even at minimal token count, skipping')
                        return
                    content = self._tokenizer.decode(
                        ids[:reduced_token_count],
                        skip_special_tokens=True,
                        clean_up_tokenization_spaces=True,
                    )
                    logger.warning(
                        f'Embedding context exceeded, retrying with {reduced_token_count} tokens'
                    )
                else:
                    raise

    async def _embed_chunks(
        self,
        chunks: list[Document],
        page_metadata: dict,
        description: str = "",
    ):
        """Embed and store pre-split chunks. Yields progress percentage (0–100).

        Uses a Semaphore to cap concurrent Ollama calls at 3.  Progress is
        reported after each chunk actually completes, not when its task is
        created (which would be immediate and misleading).
        """
        total = len(chunks)
        if total == 0:
            return

        if description:
            prefix = description[:500]
            for chunk in chunks:
                chunk.page_content = f"{prefix}\n\n{chunk.page_content}"

        sem = asyncio.Semaphore(3)

        async def _bounded(chunk: Document) -> None:
            async with sem:
                await self.process_chunk(chunk)

        tasks = [asyncio.create_task(_bounded(chunk)) for chunk in chunks]
        completed = 0
        for coro in asyncio.as_completed(tasks):
            await coro
            completed += 1
            progress = (completed / total) * 100
            if completed % 10 == 0 or completed == total:
                logger.info(
                    f'Page {page_metadata["page"]}: {completed}/{total} chunks ({progress:.0f}%)'
                )
            yield progress

    def _build_page_metadata(
        self,
        file_hash: str,
        filename: str,
        page_num: int,
        total_pages: int,
    ) -> dict:
        return {
            "source": filename,
            "filename": filename,
            "file_hash": file_hash,
            "page": page_num,
            "total_pages": total_pages,
        }

    def _load_pdf(self, file_path: Path):
        """Extrahiert Seiten aus PDF mit Tabellenerkennung und Kopf-/Fußzeilen-Filter.

        Tabellen werden via `page.find_tables()` (PyMuPDF ≥ 1.23) strukturiert als
        "Spalte: Wert | …"-Zeilen extrahiert. Tabellenregionen werden aus dem normalen
        Textextract ausgenommen, um Duplikate zu vermeiden. Kopf- und Fußzeilen werden
        via `_detect_pdf_header_footer()` erkannt und zeilenweise gefiltert.
        """
        with fitz.open(str(file_path)) as doc:
            total_pages = len(doc)
            if total_pages == 0:
                raise ValueError("Document contains no pages")
            if total_pages > _MAX_PDF_PAGES:
                raise ValueError(
                    f"PDF hat {total_pages} Seiten — Maximum: {_MAX_PDF_PAGES}. "
                    f"Limit via MAX_PDF_PAGES konfigurierbar."
                )

            header_footer = _detect_pdf_header_footer(doc)

            for i, page in enumerate(doc):
                # Tabellen erkennen und strukturiert extrahieren
                table_rects: list[fitz.Rect] = []
                table_texts: list[str] = []
                try:
                    finder = page.find_tables()
                    for tab in finder.tables:
                        table_rects.append(fitz.Rect(tab.bbox))
                        rows = tab.extract()
                        if not rows:
                            continue
                        header_row = rows[0] if rows else []
                        lines = []
                        for row in rows[1:]:
                            parts = []
                            for j, cell in enumerate(row):
                                h = str(header_row[j] or "").strip() if j < len(header_row) else ""
                                v = str(cell or "").strip()
                                if not v:
                                    continue
                                parts.append(f"{h}: {v}" if h else v)
                            if parts:
                                lines.append(" | ".join(parts))
                        if lines:
                            table_texts.append("\n".join(lines))
                except Exception:
                    pass

                if table_rects:
                    # Tabellenregionen aus Textextraktion ausschließen (verhindert Duplikate)
                    text_parts = []
                    for block in page.get_text("blocks"):
                        bx0, by0, bx1, by1, block_text, *_ = block
                        block_text = block_text.strip()
                        if not block_text:
                            continue
                        block_rect = fitz.Rect(bx0, by0, bx1, by1)
                        if any(block_rect.intersects(r) for r in table_rects):
                            continue
                        text_parts.append(block_text)
                    text = "\n".join(text_parts)
                    if table_texts:
                        text += "\n\n" + "\n\n".join(table_texts)
                else:
                    text = page.get_text()

                # Erkannte Kopf-/Fußzeilen zeilenweise entfernen
                if header_footer:
                    text = "\n".join(
                        line for line in text.split("\n")
                        if line.strip() not in header_footer
                    )

                if len(text) > _MAX_TEXT_PER_PAGE:
                    logger.warning("Seite %d: Text auf 1 MB gekürzt (%d Zeichen)", i + 1, len(text))
                    text = text[:_MAX_TEXT_PER_PAGE]

                yield text, i + 1, total_pages

    def _load_text(self, file_path: Path):
        """Lädt .txt/.md-Datei mit Encoding-Erkennung. Gibt (text, 1, 1) zurück."""
        raw = file_path.read_bytes()
        encoding = chardet.detect(raw)["encoding"] or "utf-8"
        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            text = raw.decode("utf-8", errors="replace")
            encoding = "utf-8 (Fallback)"
        safe_encodings = ("utf8", "ascii", "utf8sig")
        if encoding.lower().replace("-", "") not in safe_encodings:
            self._warnings.append(f"Als {encoding} erkannt und automatisch konvertiert")
        yield text, 1, 1

    def _load_docx(self, file_path: Path):
        """Extrahiert Text aus .docx inkl. Tabellen in Dokumentreihenfolge. Gibt (text, 1, 1) zurück."""
        from docx import Document as DocxDocument
        from docx.oxml.ns import qn

        try:
            doc = DocxDocument(file_path)
        except BadZipFile:
            raise ValueError("Datei ist passwortgeschützt oder beschädigt.")

        lines = []
        for child in doc.element.body:
            if child.tag == qn("w:p"):
                lines.append("".join(child.itertext()))
            elif child.tag == qn("w:tbl"):
                for row in child.findall(f".//{qn('w:tr')}"):
                    seen: set[int] = set()
                    cells = []
                    for tc in row.findall(qn("w:tc")):
                        if id(tc) not in seen:
                            seen.add(id(tc))
                            cells.append("".join(tc.itertext()).strip())
                    lines.append(" | ".join(c for c in cells if c))

        text = "\n".join(l for l in lines if l.strip())
        max_chars = 5_000_000
        if len(text) > max_chars:
            text = text[:max_chars]
            self._warnings.append("Dokument zu groß — Text wurde auf 5 MB begrenzt")
        yield text, 1, 1

    def _estimate_rows_per_block(self, sample_lines: list[str]) -> int:
        """Schätzt optimale Zeilen pro Embedding-Block anhand Token-Länge einer Stichprobe."""
        if not sample_lines:
            return 20
        avg_tokens = sum(len(self._tokenizer.encode(l)) for l in sample_lines) / len(sample_lines)
        return max(5, int(self.config.chunk_size * 0.8 / max(avg_tokens, 1)))

    def _load_xlsx(self, file_path: Path, sheets: list[str] | None = None):
        """Extrahiert Zeilen aus .xlsx als Blöcke. Gibt (block_text, block_num, total_blocks) zurück."""
        import openpyxl
        from datetime import datetime as dt
        from openpyxl.cell.cell import MergedCell

        wb = openpyxl.load_workbook(file_path, data_only=True)
        target_sheets = sheets if sheets else wb.sheetnames

        formula_none_count = 0
        empty_sheets = 0

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                empty_sheets += 1
                continue

            rows_iter = ws.iter_rows(values_only=False)
            header_row = next(rows_iter, None)
            if header_row is None:
                empty_sheets += 1
                continue

            headers: list[str] = []
            last_header_val: dict[int, object] = {}
            for cell in header_row:
                if isinstance(cell, MergedCell):
                    v = last_header_val.get(cell.column, "")
                else:
                    v = cell.value
                    last_header_val[cell.column] = v
                headers.append(str(v).strip() if v else f"Spalte_{cell.column}")

            all_lines: list[str] = []
            row_count = 0
            last_data_val: dict[int, object] = {}
            for row in rows_iter:
                if row_count >= self.config.max_xlsx_rows:
                    self._warnings.append(
                        f"Zeilenlimit erreicht — nur erste {self.config.max_xlsx_rows} Zeilen verarbeitet"
                    )
                    break
                row_count += 1
                parts = []
                for i, cell in enumerate(row):
                    if isinstance(cell, MergedCell):
                        v = last_data_val.get(i)
                        if v is None:
                            continue
                    else:
                        v = cell.value
                        if v is None:
                            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                formula_none_count += 1
                            continue
                        last_data_val[i] = v
                    if isinstance(v, dt):
                        v = v.strftime("%d.%m.%Y")
                    header = headers[i] if i < len(headers) else f"Spalte_{i + 1}"
                    parts.append(f"{header}: {v}")
                if parts:
                    all_lines.append(" | ".join(parts))

            rows_per_block = self._estimate_rows_per_block(all_lines[:20])
            if self.config.xlsx_max_rows_per_block is not None:
                rows_per_block = min(rows_per_block, self.config.xlsx_max_rows_per_block)
            prefix = f"Sheet: {sheet_name}\n"
            total_blocks = max(1, len(all_lines) // rows_per_block)
            for block_idx in range(0, len(all_lines), rows_per_block):
                block = prefix + "\n".join(all_lines[block_idx:block_idx + rows_per_block])
                yield block, block_idx // rows_per_block + 1, total_blocks

        if empty_sheets:
            self._warnings.append(f"{empty_sheets} leere Sheets wurden nicht verarbeitet")
        if formula_none_count:
            self._warnings.append(
                f"{formula_none_count} Zellen enthalten Formeln ohne berechneten Wert"
            )

    def _load_csv(self, file_path: Path):
        """Extrahiert Zeilen aus .csv als Blöcke. Gibt (block_text, block_num, -1) zurück."""
        import itertools

        raw = file_path.read_bytes()
        encoding = chardet.detect(raw)["encoding"] or "utf-8"
        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            text = raw.decode("utf-8", errors="replace")
            encoding = "utf-8 (Fallback)"
        safe_encodings = ("utf8", "ascii", "utf8sig")
        if encoding.lower().replace("-", "") not in safe_encodings:
            self._warnings.append(f"Als {encoding} erkannt und automatisch konvertiert")

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            reader = csv.reader(io.StringIO(text), dialect)
        except csv.Error:
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)

        first_row = next(reader, None)
        if not first_row:
            raise ValueError("CSV-Datei ist leer.")

        non_empty = [v for v in first_row if v.strip()]
        all_numeric = bool(non_empty) and all(
            v.replace(".", "").replace(",", "").replace("-", "").isdigit()
            for v in non_empty
        )
        has_header = csv.Sniffer().has_header(sample)
        if not has_header and all_numeric:
            raise ValueError("CSV ohne erkennbare Kopfzeile. Bitte Spaltennamen in Zeile 1 ergänzen.")
        headers = first_row

        # Erste 20 Zeilen puffern für Blockgrößen-Schätzung, dann replay via itertools.chain
        buffered: list[list[str]] = []
        sample_lines: list[str] = []
        for row in reader:
            line = " | ".join(f"{headers[i]}: {v}" for i, v in enumerate(row) if v.strip())
            buffered.append(row)
            sample_lines.append(line)
            if len(buffered) >= 20:
                break

        rows_per_block = self._estimate_rows_per_block(sample_lines)
        if self.config.csv_max_rows_per_block is not None:
            rows_per_block = min(rows_per_block, self.config.csv_max_rows_per_block)
        total_rows = 0
        block_lines: list[str] = []

        for row in itertools.chain(buffered, reader):
            if total_rows >= self.config.max_csv_rows:
                self._warnings.append(
                    f"Zeilenlimit erreicht — nur erste {self.config.max_csv_rows} Zeilen verarbeitet"
                )
                break
            total_rows += 1
            line = " | ".join(f"{headers[i]}: {v}" for i, v in enumerate(row) if v.strip())
            block_lines.append(line)
            if len(block_lines) >= rows_per_block:
                yield "\n".join(block_lines), total_rows // rows_per_block, -1
                block_lines = []
        if block_lines:
            yield "\n".join(block_lines), total_rows // rows_per_block + 1, -1

    async def load_documents(
        self,
        file_path: Path | str,
        original_filename: str | None = None,
        description: str = "",
        display_name: str = "",
        sheets: list[str] | None = None,
        valid_until: str | None = None,
    ) -> AsyncGenerator[float | dict, None]:
        """Load a document and store it in the vector database.

        Yields:
            float: Progress percentage (0–100)
            dict: {"already_indexed": True, "file_hash": ...} if duplicate detected
            dict: {"status": "ok", "chunk_count": N, "warnings": [...]} on completion
        """
        try:
            file_path = Path(file_path)
            fname = original_filename or file_path.name
            file_size = os.path.getsize(file_path)
            file_hash = await self._calculate_file_hash(file_path)

            existing = await self.metadata_service.get_document_metadata(file_hash)
            if existing:
                logger.info(f"Skipping already-indexed document: {fname}")
                yield {"already_indexed": True, "file_hash": file_hash}
                return

            self._warnings = []

            suffix = file_path.suffix.lower()
            if suffix == ".pdf":
                text_blocks = list(self._load_pdf(file_path))
            elif suffix in (".txt", ".md"):
                text_blocks = list(self._load_text(file_path))
            elif suffix == ".docx":
                text_blocks = list(self._load_docx(file_path))
            elif suffix == ".xlsx":
                text_blocks = list(self._load_xlsx(file_path, sheets=sheets))
            elif suffix == ".csv":
                text_blocks = list(self._load_csv(file_path))
            else:
                raise ValueError(f"Unsupported format: {suffix}")

            # Scan-Check nur für PDF (text_blocks ist bereits eine Liste, kein zweiter Pass nötig)
            if suffix == ".pdf":
                total_chars = sum(len(text) for text, _, _ in text_blocks)
                actual_total_pages = text_blocks[-1][2] if text_blocks else 0
                if total_chars < 100 and actual_total_pages > 3:
                    raise ValueError(
                        "Konnte keinen Text extrahieren. Das PDF ist möglicherweise gescannt. "
                        "Bitte eine durchsuchbare PDF-Version hochladen."
                    )

            chunk_count = 0
            total_blocks = len(text_blocks)
            for block_idx, (text, page_num, total_pages) in enumerate(text_blocks):
                page_metadata = self._build_page_metadata(file_hash, fname, page_num, total_pages)
                chunks = self._chunk_splitter.split_into_chunks(text, page_metadata)
                enhanced = self._chunk_splitter.add_neighbouring_content(chunks)
                non_empty = [c for c in enhanced if c.page_content]
                chunk_count += len(non_empty)

                block_progress = 0.0
                async for progress in self._embed_chunks(non_empty, page_metadata, description=description):
                    block_progress = progress
                    overall = ((block_idx * 100) + block_progress) / total_blocks if total_blocks > 0 else 100.0
                    yield overall

                if block_progress < 100:
                    yield ((block_idx + 1) * 100) / total_blocks if total_blocks > 0 else 100.0

            await self.metadata_service.save_document_metadata(
                file_hash,
                DocumentMetadata(
                    title=fname,
                    display_name=display_name or fname,
                    description=description,
                    valid_until=valid_until,
                    file_size=file_size,
                    page_count=total_blocks,
                    chunk_count=chunk_count,
                    source_path=str(file_path),
                    indexed_date=datetime.now(timezone.utc).isoformat(),
                    file_hash=file_hash,
                ),
            )

            yield {
                "status": "ok",
                "chunk_count": chunk_count,
                "warnings": list(self._warnings),
            }

        except Exception as e:
            logger.error(f"Error processing document {file_path}: {e}", exc_info=True)
            raise ValueError(f"Error processing document: {str(e)}") from e

    async def get_indexed_documents(self):
        """Get list of indexed documents from metadata service."""
        try:
            documents = await self.metadata_service.get_all_documents()
            return [{'title': doc.title, 'id': doc.file_hash} for doc in documents]
        except Exception as e:
            logger.error(f'Error fetching indexed documents: {e}')
            return []

    async def get_document_metadata(self, doc_id: str) -> dict | None:
        """Get document metadata by ID."""
        try:
            metadata = await self.metadata_service.get_document_metadata(doc_id)
            return metadata.model_dump() if metadata else None
        except Exception as e:
            logger.error(f'Error fetching document metadata: {e}')
            return None
